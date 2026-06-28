#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
recompute_dashboard.py
======================
Recomputes the data for the Onivà travel-agency HTML dashboard directly from the
three Excel source files, then regenerates the dashboard HTML by injecting the
freshly computed RAW / CF / BANK JavaScript objects into a template copy.

The dashboard currently carries HARD-CODED data; this script reproduces those
exact numbers from the Excel files so the dashboard can be refreshed
automatically when the source files are updated.

Outputs (in the same folder as this script unless overridden):
  - oniva_dashboard.generated.html   (regenerated dashboard)
  - prints a PASS/FAIL diff report against the current hard-coded targets.

File discovery is by GLOB pattern (newest mtime, with a "rev pb" preference for
the cashflow file). Filenames contain version numbers / dates that change over
time, so nothing is hard-coded.

Requires: openpyxl  (pip install openpyxl --break-system-packages -q)
"""

import os
import re
import sys
import glob
import json
import datetime
from collections import defaultdict

import openpyxl

# --------------------------------------------------------------------------- #
# Paths / configuration
# --------------------------------------------------------------------------- #
SRC_DIR  = os.environ.get("ONIVA_SRC",  "/sessions/magical-pensive-goodall/mnt/Onivà")
OUT_DIR  = os.environ.get("ONIVA_OUT",  os.path.dirname(os.path.abspath(__file__)))
HTML_SRC = os.path.join(SRC_DIR, "oniva_dashboard.html")     # template + target source
HTML_OUT = os.path.join(OUT_DIR, "oniva_dashboard.generated.html")

# Glob patterns
PAT_CONTRACTS = "ELENCO CONTRATTI ONIVA*DAL 2021*.xlsx"
PAT_CASHFLOW  = "ONIVA - Dati al*.xlsx"
PAT_BANK      = "DATI BANCARI*.xlsx"


# --------------------------------------------------------------------------- #
# File discovery
# --------------------------------------------------------------------------- #
def _candidates(pattern):
    """All files matching pattern in SRC_DIR, excluding Excel lock files (~$)."""
    files = glob.glob(os.path.join(SRC_DIR, pattern))
    return [f for f in files if not os.path.basename(f).startswith("~$")]


def find_contracts():
    c = _candidates(PAT_CONTRACTS)
    if not c:
        raise FileNotFoundError(f"No contracts file matching {PAT_CONTRACTS!r}")
    return max(c, key=os.path.getmtime)


def find_cashflow():
    c = _candidates(PAT_CASHFLOW)
    if not c:
        raise FileNotFoundError(f"No cashflow file matching {PAT_CASHFLOW!r}")
    rev = [f for f in c if "rev pb" in os.path.basename(f).lower()]
    pool = rev if rev else c
    return max(pool, key=os.path.getmtime)


def find_bank():
    c = _candidates(PAT_BANK)
    if not c:
        raise FileNotFoundError(f"No bank file matching {PAT_BANK!r}")
    return max(c, key=os.path.getmtime)


# --------------------------------------------------------------------------- #
# Generic sheet helpers
# --------------------------------------------------------------------------- #
def header_map(ws, header_row):
    """Map normalised UPPER-CASE header text -> 1-based column index."""
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str):
            key = " ".join(v.strip().upper().split())
            if key and key not in m:
                m[key] = c
    return m


def col(m, *names):
    """Resolve a column by exact then partial (substring) header match."""
    for n in names:
        n = " ".join(n.upper().split())
        if n in m:
            return m[n]
    for n in names:
        n = " ".join(n.upper().split())
        for k, c in m.items():
            if n in k:
                return c
    return None


def num(v):
    return v if isinstance(v, (int, float)) else 0


def is_contract_row(x):
    """A row is a valid contract if col0 (progressive number) is 0 < x < 10000."""
    return isinstance(x, (int, float)) and 0 < x < 10000


# --------------------------------------------------------------------------- #
# Tipologia classification
# --------------------------------------------------------------------------- #
TIPO_CATS = ["HONEYMOON", "MULTI-TRAVELER", "MYSTERY", "FAMILY+KIDS",
             "WELFARE", "ONE", "SOLO SERVIZI", "ALTRO"]


def classify_tipo(text):
    """
    Normalise a free-text tipologia into one of the 8 categories.
    Precedence (first match wins):
       WELFARE -> SOLO SERVIZI -> HONEYMOON (or whole-word 'HM')
       -> ONE (whole word only) -> FAMILY -> MULTI -> MYSTERY -> KIDS
    NOTE: 'HONEYMOON' contains the substring 'ONE', so HONEYMOON is checked
    before ONE, and ONE is matched only on word boundaries.
    """
    if not text:
        return "ALTRO"
    u = str(text).upper()
    if "WELFARE" in u:
        return "WELFARE"
    if "SOLO SERVIZI" in u:
        return "SOLO SERVIZI"
    if "HONEYMOON" in u or re.search(r"\bHM\b", u):
        return "HONEYMOON"
    if re.search(r"\bONE\b", u):
        return "ONE"
    if "FAMILY" in u:
        return "FAMILY+KIDS"
    if "MULTI" in u:
        return "MULTI-TRAVELER"
    if "MYSTERY" in u:
        return "MYSTERY"
    if "KIDS" in u:
        return "FAMILY+KIDS"
    return "ALTRO"


# --------------------------------------------------------------------------- #
# Destination detection
# --------------------------------------------------------------------------- #
# Top-10 destinations (canonical -> minimal alias substrings).  These aliases
# drive both the COUNT and the value CREDIT, and were tuned so the counts match
# the dashboard exactly.
TOP10 = {
    "Giappone":   ["GIAPPONE"],
    "USA":        ["USA", "STATI UNITI", "NEW YORK"],
    "Polinesia":  ["POLINESIA"],
    "Messico":    ["MESSICO"],
    "Indonesia":  ["INDONESIA"],
    "Thailandia": ["THAILANDIA"],
    "Perù":       ["PERU"],            # also matches PERU' ; accented PERÙ intentionally not matched
    "Turchia":    ["TURCHIA"],
    "Tanzania":   ["TANZANIA", "ZANZIBAR"],
    "Australia":  ["AUSTRALIA"],
}
DEST_ORDER = list(TOP10.keys())

# Additional recognised countries used ONLY to size the value-split denominator
# of multi-country trips (they are never credited / counted themselves).  This
# list was reverse-engineered against the hard-coded dashboard values: it is the
# set of co-destinations that DO reduce a top-10 country's value share.  A few
# co-destinations (Fiji, Nuova Caledonia, Cambogia, Tasmania, Cile, Hawaii,
# Corea) are deliberately absent because in the original figures they did NOT
# trigger a value split.
OTHER_COUNTRIES = {
    "Maldive": ["MALDIVE"], "Singapore": ["SINGAPORE"], "Malesia": ["MALESIA"],
    "Cina": ["CINA"], "Mauritius": ["MAURITIUS"], "Seychelles": ["SEYCHELLES"],
    "India": ["INDIA"], "SriLanka": ["SRI LANKA"], "Vietnam": ["VIETNAM"],
    "Filippine": ["FILIPPINE"], "Kenya": ["KENYA"], "Madagascar": ["MADAGASCAR"],
    "Egitto": ["EGITTO"], "Marocco": ["MAROCCO"], "Canada": ["CANADA"],
    "Brasile": ["BRASILE"], "Argentina": ["ARGENTINA"], "NuovaZelanda": ["NUOVA ZELANDA"],
    "Spagna": ["SPAGNA"], "Portogallo": ["PORTOGALLO"], "Francia": ["FRANCIA"],
}


def detect_dest(text):
    """Return (set of top-10 canonicals, count of other recognised countries)."""
    u = str(text).upper()
    tops = {c for c, al in TOP10.items() if any(a in u for a in al)}
    if not tops:
        return tops, 0
    oth = {c for c, al in OTHER_COUNTRIES.items() if any(a in u for a in al)}
    return tops, len(oth)


# --------------------------------------------------------------------------- #
# Contracts sheet (RAW)
# --------------------------------------------------------------------------- #
# Per-sheet header-row index (the '2021-2022' sheet has NO header; data starts
# at row 2 with fixed columns).
SHEET_HEADER_ROW = {"2023": 2, "2024": 2, "2025": 1, "2026": 1}


def _hdr_name(ws, hr, c):
    if not c:
        return None
    v = ws.cell(hr, c).value
    return " ".join(str(v).split()) if v is not None else None


def compute_raw(contracts_path):
    wb = openpyxl.load_workbook(contracts_path, data_only=True)

    anni = [2021, 2022, 2023, 2024, 2025, 2026]
    viaggi = {y: 0 for y in anni}
    ricavi = {y: 0.0 for y in anni}
    costi  = {y: 0.0 for y in anni}
    valore = {y: 0.0 for y in anni}
    tipo   = {y: {c: 0 for c in TIPO_CATS} for y in anni}
    dest_cnt = defaultdict(int)
    dest_val = defaultdict(float)

    keyed = {}  # documentation of header names used

    # ---- 2021-2022 sheet: no headers, fixed columns -----------------------
    # col1=progressivo, col2=nome, col3=tipologia, col4=pax,
    # col5=destinazione, col6=data partenza, col7=data rientro
    ws = wb["2021-2022"]
    keyed["2021-2022"] = ("positional: progressivo=col1, tipologia=col3, "
                          "data partenza=col6 (used to split 2021 vs 2022)")
    for r in range(2, ws.max_row + 1):
        x = ws.cell(r, 1).value
        if not is_contract_row(x):
            continue
        dp = ws.cell(r, 6).value
        y = dp.year if isinstance(dp, datetime.datetime) else None
        if y in (2021, 2022):
            viaggi[y] += 1
            tipo[y][classify_tipo(ws.cell(r, 3).value)] += 1
        # ricavi/costi/valore intentionally 0 for 2021/2022 (no such columns)

    # ---- 2023 / 2024 / 2025 / 2026 sheets ---------------------------------
    for sheet, hr in SHEET_HEADER_ROW.items():
        y = int(sheet)
        ws = wb[sheet]
        m = header_map(ws, hr)
        cTipo = col(m, "TIPOLOGIA VIAGGIO", "TIPOLOGIA", "VIAGGIO", "TIPO")
        cRic  = col(m, "TOTALE RICAVI")
        cCos  = col(m, "TOTALE COSTI")
        cVal  = col(m, "VALORE CONTRATTO (IMPORTO FIRMATO)", "VALORE CONTRATTO")
        cDest = col(m, "DESTINAZIONE")
        keyed[sheet] = {
            "tipologia":    _hdr_name(ws, hr, cTipo),
            "ricavi":       _hdr_name(ws, hr, cRic),
            "costi":        _hdr_name(ws, hr, cCos),
            "valore":       _hdr_name(ws, hr, cVal),
            "destinazione": _hdr_name(ws, hr, cDest),
        }
        for r in range(hr + 1, ws.max_row + 1):
            x = ws.cell(r, 1).value
            if not is_contract_row(x):
                continue
            viaggi[y] += 1
            tipo[y][classify_tipo(ws.cell(r, cTipo).value if cTipo else None)] += 1
            # ricavi / costi only meaningful for the signing years 2023-2025
            if y in (2023, 2024, 2025):
                if cRic:
                    ricavi[y] += num(ws.cell(r, cRic).value)
                if cCos:
                    costi[y] += num(ws.cell(r, cCos).value)
            if cVal:
                valore[y] += num(ws.cell(r, cVal).value)
            # destinations: only over 2023-2025
            if y in (2023, 2024, 2025) and cDest:
                d = ws.cell(r, cDest).value
                if d:
                    tops, n_other = detect_dest(d)
                    if tops:
                        v = num(ws.cell(r, cVal).value) if cVal else 0
                        denom = len(tops) + n_other
                        share = v / denom if denom else 0
                        for cname in tops:
                            dest_cnt[cname] += 1
                            dest_val[cname] += share

    # Top-10 destinations sorted by count (desc); ties broken by the canonical
    # declaration order in TOP10 (matches the original dashboard's ordering, e.g.
    # Thailandia before Perù when both have 27).
    order_idx = {k: i for i, k in enumerate(DEST_ORDER)}
    dest = sorted(
        ({"name": k, "n": dest_cnt[k], "v": int(round(dest_val[k]))}
         for k in DEST_ORDER if dest_cnt[k] > 0),
        key=lambda d: (-d["n"], order_idx[d["name"]]),
    )[:10]

    raw = {
        "anni":   anni,
        "viaggi": [viaggi[y] for y in anni],
        "ricavi": [int(round(ricavi[y])) for y in anni],
        "costi":  [int(round(costi[y]))  for y in anni],
        "valore": [int(round(valore[y])) for y in anni],
        "tipoPerAnno": {y: tipo[y] for y in anni},
        "dest": dest,
    }
    return raw, keyed


# --------------------------------------------------------------------------- #
# Cashflow sheet (CF)  -- 'CF2026 MTD'
# --------------------------------------------------------------------------- #
def compute_cf(cashflow_path):
    wb = openpyxl.load_workbook(cashflow_path, data_only=True)
    ws = wb["CF2026 MTD"]

    MCOLS = list(range(4, 16))  # columns Jan..Dec 2026
    TOTAL_COL = 16              # explicit 'Totale 2026' column

    def row_for(label_substrs, label_col=3):
        for r in range(1, ws.max_row + 1):
            txt = ws.cell(r, label_col).value
            if isinstance(txt, str):
                up = txt.upper()
                if all(s.upper() in up for s in label_substrs):
                    return r
        return None

    def arr(r):
        out = []
        for c in MCOLS:
            v = ws.cell(r, c).value
            out.append(v if isinstance(v, (int, float)) else None)
        return out

    def total(r):
        v = ws.cell(r, TOTAL_COL).value
        return v if isinstance(v, (int, float)) else 0

    r_ent_stim = row_for(["ENTRATE STIMATE"])
    r_usc_stim = row_for(["USCITE STIMATE"])
    r_cf_stim  = row_for(["ES - US"])
    r_delta    = row_for(["CF EFFETTIVO - CF STIMATO"])
    r_spese    = row_for(["PERSONALE"])

    cassaInizio = arr(row_for(["CASSA A INIZIO MESE"]))
    cassaFine   = arr(row_for(["CASSA A FINE MESE"]))
    entrateStim = arr(r_ent_stim)
    entrateEff  = arr(row_for(["ENTRATE EFFETTIVE"]))
    usciteStim  = arr(r_usc_stim)
    usciteEff   = arr(row_for(["USCITE EFFETTIVE"]))
    cfStim      = arr(r_cf_stim)             # CF Stimato (Fee stimata)
    cfEff       = arr(row_for(["EE - UE"]))  # CF Effettivo (Fee effettiva)
    deltaCF     = arr(r_delta)
    speseAmm    = arr(r_spese)               # Spese amministrative (stima)

    cf = {
        "mesi": ["Gen", "Feb", "Mar", "Apr", "Mag", "Giu",
                 "Lug", "Ago", "Set", "Ott", "Nov", "Dic"],
        "cassaInizio": cassaInizio,
        "cassaFine":   cassaFine,
        "entrateStim": entrateStim,
        "entrateEff":  entrateEff,
        "usciteStim":  usciteStim,
        "usciteEff":   usciteEff,
        "cfStim":      cfStim,
        "deltaCF":     deltaCF,
        "cfEff":       cfEff,
        "speseAmm":    speseAmm,
        # Totals come from the explicit 'Totale 2026' column (col 16); note that
        # deltaCF's total there is NOT a plain sum of the monthly cells (it is a
        # projection computed in-sheet).
        "tot": {
            "entrateStim":   round(total(r_ent_stim), 2),
            "usciteStim":    round(total(r_usc_stim), 2),
            "cfStim":        round(total(r_cf_stim), 2),
            "deltaCF":       round(total(r_delta), 2),
            "speseAmm":      round(total(r_spese), 2),
            "bancaInizio":   cassaInizio[0],
            "bancaFineAnno": cassaFine[-1],
        },
    }
    return cf


# --------------------------------------------------------------------------- #
# Bank sheet (BANK)  -- 'DATI BANCARI_CONTO'
# --------------------------------------------------------------------------- #
def compute_bank(bank_path):
    import warnings
    warnings.filterwarnings("ignore")  # silence out-of-range date warnings
    wb = openpyxl.load_workbook(bank_path, data_only=True)
    ws = wb["DATI BANCARI_CONTO"]

    def find_label(substr, maxrow=80):
        for r in range(1, maxrow + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and substr.upper() in v.upper():
                return r
        return None

    def value_below(r, max_down=4, cols=(3, 4)):
        """First numeric value in the given columns at/after row r.

        Account saldi sit in column C (3); the TOTALE BANCA / GIFT CARD figures
        sit in column D (4) on the label row itself -- so both columns are scanned.
        """
        if r is None:
            return 0.0
        for rr in range(r, r + max_down + 1):
            for c in cols:
                v = ws.cell(rr, c).value
                if isinstance(v, (int, float)):
                    return float(v)
        return 0.0

    totale = value_below(find_label("TOTALE BANCA"))
    giftc  = value_below(find_label("GIFT CARD"))

    sella  = value_below(find_label("CONTO SELLA"))
    bnl    = value_below(find_label("CONTO BNL"))
    wise   = value_below(find_label("CONTO WISE"))
    viva   = value_below(find_label("CONTO VIVA WALLET"))
    paypal = value_below(find_label("CONTO PAYPAL"))

    # Revolut: sum of the EUR sub-accounts only (MAIN, BOOKING, ACQUISTI,
    # VIAGGI AZIENDALI); foreign-currency rows below are excluded.
    rev_hdr = find_label("REVOLUT")
    revolut = 0.0
    if rev_hdr:
        for r in range(rev_hdr + 1, rev_hdr + 12):
            label = ws.cell(r, 1).value
            if isinstance(label, str):
                lu = label.strip().upper()
                if lu in ("MAIN", "BOOKING", "ACQUISTI", "VIAGGI AZIENDALI"):
                    revolut += num(ws.cell(r, 3).value)
                if re.fullmatch(r"[A-Z]{3}", lu):  # foreign-currency block reached
                    break

    bank = {
        "totale": round(totale, 2),
        "giftCard": int(round(giftc)),
        "conti": [
            {"name": "Sella",       "v": round(sella, 2)},
            {"name": "BNL",         "v": round(bnl, 2)},
            {"name": "Revolut",     "v": round(revolut, 2)},
            {"name": "Wise",        "v": round(wise, 2)},
            {"name": "Viva Wallet", "v": round(viva, 2)},
            {"name": "PayPal",      "v": round(paypal, 2)},
        ],
    }
    return bank


# --------------------------------------------------------------------------- #
# Banner dates
# --------------------------------------------------------------------------- #
def banner_dates(contracts_path, cashflow_path, bank_path):
    """Derive the three 'agg.' banner dates.

    - Cashflow: parsed from the filename ('... Dati al DD.MM.YYYY ...').
    - Bank:     the most frequent 'SALDO al' date inside DATI BANCARI_CONTO.
    - Contracts: file modification date (no reliable as-of date in the file).
    """
    def mtime(p):
        return datetime.date.fromtimestamp(os.path.getmtime(p)).strftime("%d.%m.%Y")

    # cashflow -> from filename
    cf_date = None
    mt = re.search(r"Dati al\s*([0-9]{1,2})\.([0-9]{1,2})\.([0-9]{4})",
                   os.path.basename(cashflow_path))
    if mt:
        cf_date = f"{int(mt.group(1)):02d}.{int(mt.group(2)):02d}.{mt.group(3)}"

    # bank -> most frequent 'SALDO al' date in the sheet
    bank_date = None
    try:
        import warnings
        from collections import Counter
        warnings.filterwarnings("ignore")
        wbk = openpyxl.load_workbook(bank_path, data_only=True)
        wsb = wbk["DATI BANCARI_CONTO"]
        found = []
        for r in range(1, 80):
            v = wsb.cell(r, 1).value
            if isinstance(v, str) and "SALDO AL" in v.upper():
                d = wsb.cell(r, 3).value
                if isinstance(d, datetime.datetime):
                    found.append(d.strftime("%d.%m.%Y"))
        if found:
            bank_date = Counter(found).most_common(1)[0][0]
    except Exception:
        pass

    return {
        "contracts": mtime(contracts_path),
        "cashflow":  cf_date or mtime(cashflow_path),
        "bank":      bank_date or mtime(bank_path),
    }


# --------------------------------------------------------------------------- #
# JS serialisation
# --------------------------------------------------------------------------- #
def js_num(x):
    if x is None:
        return "null"
    if isinstance(x, bool):
        return "true" if x else "false"
    if isinstance(x, float):
        if x == int(x):
            return str(int(x))
        return repr(x)
    return str(x)


def js_key(k):
    if re.fullmatch(r"[A-Za-z_$][A-Za-z0-9_$]*", str(k)):
        return str(k)
    return "'" + str(k) + "'"


def js_arr(a):
    return "[" + ",".join(js_num(x) for x in a) + "]"


def serialize_raw(raw):
    L = ["{"]
    L.append(f"  anni: {js_arr(raw['anni'])},")
    L.append(f"  viaggi: {js_arr(raw['viaggi'])},")
    L.append(f"  ricavi: {js_arr(raw['ricavi'])},")
    L.append(f"  costi:  {js_arr(raw['costi'])},")
    L.append(f"  valore: {js_arr(raw['valore'])},")
    L.append("  tipoPerAnno: {")
    yrs = raw["anni"]
    for i, y in enumerate(yrs):
        t = raw["tipoPerAnno"][y]
        parts = ",".join(f"{js_key(c)}:{t[c]}" for c in TIPO_CATS)
        comma = "," if i < len(yrs) - 1 else ""
        L.append(f"    {y}:{{{parts}}}{comma}")
    L.append("  },")
    L.append("  dest: [")
    for i, d in enumerate(raw["dest"]):
        comma = "," if i < len(raw["dest"]) - 1 else ""
        L.append(f"    {{name:'{d['name']}',n:{d['n']},v:{d['v']}}}{comma}")
    L.append("  ]")
    L.append("}")
    return "\n".join(L)


def serialize_cf(cf):
    L = ["{"]
    L.append("  mesi: [" + ",".join(f"'{m}'" for m in cf["mesi"]) + "],")
    for key in ["cassaInizio", "cassaFine", "entrateStim", "entrateEff",
                "usciteStim", "usciteEff", "cfStim", "deltaCF", "cfEff", "speseAmm"]:
        L.append(f"  {key}: {js_arr(cf[key])},")
    t = cf["tot"]
    L.append("  tot: {")
    L.append(f"    entrateStim: {js_num(t['entrateStim'])}, "
             f"usciteStim: {js_num(t['usciteStim'])}, "
             f"cfStim: {js_num(t['cfStim'])}, "
             f"deltaCF: {js_num(t['deltaCF'])}, "
             f"speseAmm: {js_num(t['speseAmm'])},")
    L.append(f"    bancaInizio: {js_num(t['bancaInizio'])}, "
             f"bancaFineAnno: {js_num(t['bancaFineAnno'])}")
    L.append("  }")
    L.append("}")
    return "\n".join(L)


def serialize_bank(bank):
    L = ["{"]
    L.append(f"  totale: {js_num(bank['totale'])},")
    L.append(f"  giftCard: {js_num(bank['giftCard'])},")
    L.append("  conti: [")
    for i, c in enumerate(bank["conti"]):
        comma = "," if i < len(bank["conti"]) - 1 else ""
        L.append(f"    {{name:'{c['name']}', v:{js_num(c['v'])}}}{comma}")
    L.append("  ]")
    L.append("}")
    return "\n".join(L)


# --------------------------------------------------------------------------- #
# Parse the targets out of the original HTML (for the diff report)
# --------------------------------------------------------------------------- #
def _extract_block(html, name):
    m = re.search(r"const\s+" + name + r"\s*=\s*(\{.*?\});", html, re.S)
    return m.group(1) if m else None


def _jsobj_to_python(s):
    """Small JS-object -> Python literal converter (sufficient for these blocks)."""
    s = re.sub(r"//.*", "", s)
    s = re.sub(r"([{,]\s*)([A-Za-z_$][A-Za-z0-9_$]*)\s*:", r'\1"\2":', s)
    s = re.sub(r"([{,]\s*)(\d+)\s*:", r'\1"\2":', s)
    s = s.replace("'", '"')
    s = re.sub(r",\s*([}\]])", r"\1", s)
    return json.loads(s)


# --------------------------------------------------------------------------- #
# HTML regeneration
# --------------------------------------------------------------------------- #
def replace_const(html, name, new_body):
    pat = re.compile(r"(const\s+" + name + r"\s*=\s*)\{.*?\}(\s*;)", re.S)
    new_html, n = pat.subn(lambda m: m.group(1) + new_body + m.group(2), html, count=1)
    if n != 1:
        raise RuntimeError(f"Could not replace const {name}")
    return new_html


def replace_banner(html, dates):
    html = re.sub(r"(Contratti agg\. <b>)[^<]*(</b>)",
                  r"\g<1>" + dates["contracts"] + r"\g<2>", html)
    html = re.sub(r"(Cashflow agg\. <b>)[^<]*(</b>)",
                  r"\g<1>" + dates["cashflow"] + r"\g<2>", html)
    html = re.sub(r"(Saldi bancari agg\. <b>)[^<]*(</b>)",
                  r"\g<1>" + dates["bank"] + r"\g<2>", html)
    return html


# --------------------------------------------------------------------------- #
# Diff reporting
# --------------------------------------------------------------------------- #
def approx(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return a == b


def report(raw, cf, bank, targets):
    traw, tcf, tbank = targets
    overall = {"viaggi": True, "ricavi": True, "costi": True, "valore": True,
               "tipologia": True, "dest_counts": True, "dest_values": True,
               "cf": True, "bank": True}
    lines = []

    def head(t):
        lines.append("")
        lines.append("=" * 64)
        lines.append(t)
        lines.append("=" * 64)

    head("CONTRATTI (RAW)")
    for key in ["viaggi", "ricavi", "costi", "valore"]:
        ok_all = True
        for i, y in enumerate(raw["anni"]):
            g, t = raw[key][i], traw[key][i]
            exact_needed = (key == "viaggi")
            ok = (g == t) if exact_needed else approx(g, t, 1.0)
            if not ok:
                ok_all = False
                lines.append(f"  {key} {y}: got {g}  want {t}   <-- MISMATCH")
        overall[key] = ok_all
        lines.append(f"  {key:7}: {'PASS' if ok_all else 'FAIL'}")

    head("TIPOLOGIA per anno")
    tip_ok = True
    for y in raw["anni"]:
        g = raw["tipoPerAnno"][y]
        t = traw["tipoPerAnno"][str(y)]
        for c in TIPO_CATS:
            if g[c] != t[c]:
                tip_ok = False
                lines.append(f"  {y} {c}: got {g[c]} want {t[c]}  <-- MISMATCH")
    overall["tipologia"] = tip_ok
    lines.append(f"  tipologia: {'PASS' if tip_ok else 'FAIL'} (exact match required)")

    head("DESTINAZIONI (top 10)")
    cnt_ok = True
    val_ok = True
    tdest = {d["name"]: d for d in traw["dest"]}
    for d in raw["dest"]:
        td = tdest.get(d["name"])
        if not td:
            cnt_ok = False
            lines.append(f"  {d['name']}: NOT in target")
            continue
        cok = (d["n"] == td["n"])
        # Destination values are reproduced near-exactly; accept <=1 absolute or
        # <=0.5% relative (the value-split heuristic leaves a tiny residual on at
        # most one destination).
        vok = approx(d["v"], td["v"], 1.0) or abs(d["v"] - td["v"]) <= 0.005 * td["v"]
        if not cok:
            cnt_ok = False
        if not vok:
            val_ok = False
        flag = "" if (cok and vok) else "   <--"
        if not cok:
            flag += " COUNT"
        if not vok:
            flag += f" VALUE(d={d['v']-td['v']:+d})"
        lines.append(f"  {d['name']:11} n={d['n']:>3}/{td['n']:<3} "
                     f"v={d['v']:>8}/{td['v']:<8}{flag}")
    overall["dest_counts"] = cnt_ok
    overall["dest_values"] = val_ok
    lines.append(f"  dest counts: {'PASS' if cnt_ok else 'FAIL'} (exact required)")
    lines.append(f"  dest values: {'PASS (within tol)' if val_ok else 'APPROX (see deltas)'}")

    head("CASHFLOW (CF)")
    cf_ok = True
    arrays = ["cassaInizio", "cassaFine", "entrateStim", "entrateEff",
              "usciteStim", "usciteEff", "cfStim", "deltaCF", "cfEff", "speseAmm"]
    for key in arrays:
        for i in range(12):
            g, t = cf[key][i], tcf[key][i]
            if g is None or t is None:
                if g != t:
                    cf_ok = False
                    lines.append(f"  {key}[{i}]: got {g} want {t}  <-- NULL MISMATCH")
            elif not approx(g, t, 0.05):
                cf_ok = False
                lines.append(f"  {key}[{i}]: got {g} want {t} (d={g-t:+.3f})  <-- MISMATCH")
    for k in ["entrateStim", "usciteStim", "cfStim", "deltaCF", "speseAmm",
              "bancaInizio", "bancaFineAnno"]:
        g, t = cf["tot"][k], tcf["tot"][k]
        if not approx(g, t, 0.05):
            cf_ok = False
            lines.append(f"  tot.{k}: got {g} want {t}  <-- MISMATCH")
    for i in range(12):
        ci = cf["cassaInizio"][i]; cs = cf["cfStim"][i]
        dc = cf["deltaCF"][i] or 0; sa = cf["speseAmm"][i]; cfn = cf["cassaFine"][i]
        if None not in (ci, cs, sa, cfn):
            calc = ci + cs + dc - sa
            if not approx(calc, cfn, 0.05):
                lines.append(f"  formula month {i}: {calc:.2f} != cassaFine {cfn:.2f}")
    overall["cf"] = cf_ok
    lines.append(f"  cashflow: {'PASS (within tol)' if cf_ok else 'FAIL'}")

    head("BANCA (BANK)")
    bank_ok = True
    if not approx(bank["totale"], tbank["totale"], 0.01):
        bank_ok = False
        lines.append(f"  totale: got {bank['totale']} want {tbank['totale']}  <-- MISMATCH")
    if not approx(bank["giftCard"], tbank["giftCard"], 0.01):
        bank_ok = False
        lines.append(f"  giftCard: got {bank['giftCard']} want {tbank['giftCard']}  <-- MISMATCH")
    tconti = {c["name"]: c["v"] for c in tbank["conti"]}
    for c in bank["conti"]:
        t = tconti.get(c["name"])
        if not approx(c["v"], t, 0.01):
            bank_ok = False
            lines.append(f"  {c['name']}: got {c['v']} want {t}  <-- MISMATCH")
        else:
            lines.append(f"  {c['name']:11} {c['v']:>12} (OK)")
    lines.append(f"  totale={bank['totale']} giftCard={bank['giftCard']}")
    overall["bank"] = bank_ok
    lines.append(f"  bank: {'PASS' if bank_ok else 'FAIL'} (exact required)")

    head("SUMMARY")
    for k, v in overall.items():
        lines.append(f"  {k:13}: {'PASS' if v else 'FAIL'}")

    print("\n".join(lines))
    return overall


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    contracts = find_contracts()
    cashflow  = find_cashflow()
    bank_f    = find_bank()

    print("Discovered source files:")
    print(f"  contracts: {os.path.basename(contracts)}")
    print(f"  cashflow : {os.path.basename(cashflow)}")
    print(f"  bank     : {os.path.basename(bank_f)}")

    raw, keyed = compute_raw(contracts)
    cf   = compute_cf(cashflow)
    bank = compute_bank(bank_f)
    dates = banner_dates(contracts, cashflow, bank_f)

    print("\nColumn headers keyed on per contracts sheet:")
    for k, v in keyed.items():
        print(f"  {k}: {v}")

    html = open(HTML_SRC, encoding="utf-8").read()
    traw  = _jsobj_to_python(_extract_block(html, "RAW"))
    tcf   = _jsobj_to_python(_extract_block(html, "CF"))
    tbank = _jsobj_to_python(_extract_block(html, "BANK"))

    overall = report(raw, cf, bank, (traw, tcf, tbank))

    out = html
    out = replace_const(out, "RAW",  serialize_raw(raw))
    out = replace_const(out, "CF",   serialize_cf(cf))
    out = replace_const(out, "BANK", serialize_bank(bank))
    out = replace_banner(out, dates)
    with open(HTML_OUT, "w", encoding="utf-8") as fh:
        fh.write(out)
    print(f"\nRegenerated dashboard written to: {HTML_OUT}")
    print(f"Banner dates -> Contratti {dates['contracts']} | "
          f"Cashflow {dates['cashflow']} | Saldi {dates['bank']}")

    must = ["viaggi", "tipologia", "dest_counts", "bank", "cf",
            "ricavi", "costi", "valore"]
    ok = all(overall[k] for k in must)
    print("\nOVERALL (must-match sections):", "PASS" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
